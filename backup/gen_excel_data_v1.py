import datetime
import os.path
from log_config import setup_logging
import logging

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from get_excel_data_curr.ConfigTool import ConfigTool

def convert_building_show(bid):
    if bid == 11:
        return '11A'
    elif bid == 12:
        return '11B'
    elif bid == 13:
        return '12A'
    elif bid == 14:
        return '12B'
    else:
        return str(bid)

def gen_excel_data_v1(ret_dict, username):
    if len(ret_dict) == 0:
        print('ret_dict is None')
        return

    # 创建一个新的工作簿
    new_workbook = Workbook()

    all_data = []

    for bid, ret_data in ret_dict.items():
        for row in ret_data:
            row['roomName'] = '{}-{}'.format(convert_building_show(int(bid)), str(row['roomName']))

    for bid, ret_data in ret_dict.items():
        all_data += ret_data

    new_sheet1 = new_workbook.active

    new_sheet1.title = '-'.join(convert_building_show(int(bid)) for bid in ret_dict.keys())

    new_sheet1['A1'] = '日期'
    new_sheet1['B1'] = '学院'
    new_sheet1['C1'] = '学号'
    new_sheet1['D1'] = '姓名'
    new_sheet1['E1'] = '宿舍号'
    new_sheet1['F1'] = '年级'
    new_sheet1['G1'] = '培养层次'
    new_sheet1['H1'] = '晚归时间'

    idx = 2
    for row in all_data:
        # 日期列
        date_time = datetime.datetime.strptime(str(row['passTimeText']), "%Y-%m-%d %H:%M:%S")
        # 格式化输出为"月.日"的形式
        formatted_date = f"{date_time.month}.{date_time.day}"
        new_sheet1[f'A{idx}'] = formatted_date

        # 学院列
        start = row['schoolInstituteName'][0:2]  # 开头的字符
        end = row['schoolInstituteName'][-2:]  # 结尾的字符
        config_tool = ConfigTool("./get_excel_data_curr/config.json")
        data_cfg = config_tool.get_data_cfg()
        if row['schoolInstituteName'] in data_cfg:
            new_value = data_cfg[row['schoolInstituteName']]
        else:
            new_value = start + end  # 拼接新的值
        new_sheet1[f'B{idx}'] = new_value

        # 学号列
        new_sheet1[f'C{idx}'] = str(row['userId'])

        # 姓名列
        new_sheet1[f'D{idx}'] = str(row['userName'])

        # 宿舍号列
        new_val = str(row['roomName'])
        new_sheet1[f'E{idx}'] = new_val

        # 年级列
        new_sheet1[f'F{idx}'] = str(row['grade'])

        # 培养层次
        new_sheet1[f'G{idx}'] = '研究生'

        # 晚归时间列
        new_sheet1[f'H{idx}'] = str(row['passTimeText'])[10:16]

        # 设置样式
        set_style(new_sheet1)

        idx += 1

    ## 保存修改后的工作簿
    # 获取当前时间
    now = datetime.datetime.now()
    # 减去一天
    one_day = datetime.timedelta(days=1)
    yesterday = now - one_day

    # 格式化昨天的时间为月日的形式，例如 9.19
    formatted_yesterday = f"{yesterday.month}.{yesterday.day}"

    path = f"/result-files/{username}"
    if not os.path.exists(path):
        os.makedirs(path)

    file_name = '公寓学生晚归名单{}.xlsx'.format(formatted_yesterday)
    file_path = f'.{path}/{file_name}'
    if os.path.exists(file_path):
        os.remove(file_path)
        print('======================旧数据删除完成========================')
    new_workbook.save(file_path)
    print('======================数据导出完成========================')
    return file_path


def set_style(new_sheet1):
    # 设置单元格对齐样式
    align = Alignment(horizontal='center', vertical='center')
    # 设置单元格边框样式
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    # 遍历工作表中的所有单元格并设置样式
    for row in new_sheet1.iter_rows():
        for cell in row:
            cell.alignment = align
            cell.border = border
    # 设置自适应列宽
    for col in new_sheet1.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # 获取列字母
        for cell in col:
            try:  # 需要尝试因为如果单元格为空将会导致错误
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 2
        new_sheet1.column_dimensions[column].width = adjusted_width
