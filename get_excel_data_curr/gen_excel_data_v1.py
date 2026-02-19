import datetime
import os.path
from log_config import setup_logging
import logging

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
from operator import itemgetter

setup_logging()  # 多次调用不会重复配置

def process_data(data):
    # 创建一个字典，用于存储最终结果
    result = {}

    # 遍历外层的字典
    for key, records in data.items():
        # 创建一个字典，用于存储按userId分组的数据
        grouped_data = defaultdict(list)

        # 遍历每个记录
        for record in records:
            # 按userId将记录添加到相应的分组中
            grouped_data[record['userId']].append(record)

        # 对每个userId的记录按passTimeText降序排序，并保留最新的记录
        all_data = []
        for user_id, records in grouped_data.items():
            # 按passTimeText降序排序
            records.sort(key=itemgetter('passTimeText'), reverse=True)
            # 保留最新的记录
            latest_record = records[0]
            # 去重，只保留passTimeText最大的一条数据
            grouped_data[user_id] = [latest_record]
            all_data.append(latest_record)

        # 将处理后的数据添加到结果字典中，保留外层的key
        result[key] = all_data

    return result

def convert_building_show(bid):
    if bid == 11:
        return '11A'
    elif bid == 12:
        return '11B'
    elif bid == 13:
        return '12A'
    elif bid == 14:
        return '12B'
    elif bid == 15:
        return '13'
    elif bid == 16:
        return '14'
    elif bid == 17:
        return '15'
    elif bid == 18:
        return '16'
    elif bid == 19:
        return '17A'
    elif bid == 20:
        return '17B'
    else:
        return str(bid)

def gen_excel_data_v1(ret_dict, username, data_cfg=None, request_data=None):
    """
    生成 Excel 晚归数据文件。
    :param ret_dict: 各楼栋的晚归数据
    :param username: 用户名（用于文件路径）
    :param data_cfg: 学院名称映射字典，从调用方传入
    :param request_data: 请求数据（包含 startDate, endDate 等）
    """
    if data_cfg is None:
        data_cfg = {}
    if len(ret_dict) == 0:
        logging.debug('ret_dict is None')
        return

    # 创建一个新的工作簿
    new_workbook = Workbook()

    all_data = []

    for bid, ret_data in ret_dict.items():
        for row in ret_data:
            row['roomName'] = '{}-{}'.format(convert_building_show(int(bid)), str(row['roomName']))

    # 调用函数处理数据
    ret_dict_new = process_data(ret_dict)

    for bid, ret_data in ret_dict_new.items():
        all_data += ret_data

    # 排序：先按学院升序，再按日期升序，最后按晚归时间升序
    def sort_key(row):
        # 学院：使用映射后的名称排序
        institute = row.get('schoolInstituteName', '')
        if institute in data_cfg:
            institute_name = data_cfg[institute]
        else:
            institute_name = institute[0:2] + institute[-2:] if len(institute) >= 2 else institute
        # 日期和时间直接用 passTimeText 字符串排序（格式 YYYY-MM-DD HH:MM:SS，天然支持字典序）
        pass_time = row.get('passTimeText', '')
        return (institute_name, pass_time)

    all_data.sort(key=sort_key)

    new_sheet1 = new_workbook.active

    # 生成sheet名称，Excel sheet名称最多31个字符
    sheet_title = '-'.join(convert_building_show(int(bid)) for bid in ret_dict.keys())
    if len(sheet_title) > 31:
        # 名称过长时，显示第一个和最后一个楼栋，中间用省略号
        building_list = [convert_building_show(int(bid)) for bid in ret_dict.keys()]
        sheet_title = f"{building_list[0]}至{building_list[-1]}栋({len(building_list)}栋)"
        if len(sheet_title) > 31:
            sheet_title = sheet_title[:31]
    new_sheet1.title = sheet_title

    new_sheet1['A1'] = '日期'
    new_sheet1['B1'] = '学院'
    new_sheet1['C1'] = '学号'
    new_sheet1['D1'] = '姓名'
    new_sheet1['E1'] = '宿舍号'
    new_sheet1['F1'] = '年级'
    new_sheet1['G1'] = '培养层次'
    new_sheet1['H1'] = '晚归时间'

    idx = 2
    for row in all_data:
        # 日期列
        date_time = datetime.datetime.strptime(str(row['passTimeText']), "%Y-%m-%d %H:%M:%S")
        # 格式化输出为"月.日"的形式
        formatted_date = f"{date_time.month}.{date_time.day}"
        new_sheet1[f'A{idx}'] = formatted_date

        # 学院列
        start = row['schoolInstituteName'][0:2]  # 开头的字符
        end = row['schoolInstituteName'][-2:]  # 结尾的字符
        if row['schoolInstituteName'] in data_cfg:
            new_value = data_cfg[row['schoolInstituteName']]
        else:
            new_value = start + end  # 拼接新的值
        new_sheet1[f'B{idx}'] = new_value

        # 学号列
        new_sheet1[f'C{idx}'] = str(row['userId'])

        # 姓名列
        new_sheet1[f'D{idx}'] = str(row['userName'])

        # 宿舍号列
        new_val = str(row['roomName'])
        new_sheet1[f'E{idx}'] = new_val

        # 年级列
        new_sheet1[f'F{idx}'] = str(row['grade'])

        # 培养层次：根据楼栋和宿舍号判断
        # 4、5、7、11A、11B栋是研究生，其他是本科
        # 4栋部分宿舍是本科
        room_name = str(row['roomName'])
        building_num = room_name.split('-')[0] if '-' in room_name else ''

        # 4栋本科宿舍列表
        building_4_undergraduate = [
            '4-101', '4-103', '4-105', '4-107', '4-109', '4-111', '4-113', '4-115', '4-117',
            '4-121', '4-123', '4-125', '4-127', '4-129', '4-137', '4-139', '4-141', '4-143',
            '4-201', '4-202', '4-203', '4-204', '4-205', '4-206', '4-207', '4-208', '4-209', '4-210',
            '4-211', '4-212', '4-213', '4-214', '4-215', '4-216', '4-217', '4-218', '4-219', '4-220',
            '4-221', '4-222', '4-223', '4-224', '4-225', '4-227', '4-228'
        ]

        if room_name in building_4_undergraduate:
            new_sheet1[f'G{idx}'] = '本科'
        elif building_num in ['4', '5', '7', '11A', '11B']:
            new_sheet1[f'G{idx}'] = '研究生'
        else:
            new_sheet1[f'G{idx}'] = '本科'

        # 晚归时间列
        new_sheet1[f'H{idx}'] = str(row['passTimeText'])[10:16]

        # 设置样式
        set_style(new_sheet1)

        idx += 1

    ## 保存修改后的工作簿
    # 根据查询的起止日期生成文件名
    if request_data and request_data.get('startDate') and request_data.get('endDate'):
        start_dt = datetime.datetime.strptime(request_data['startDate'], "%Y-%m-%d")
        end_dt = datetime.datetime.strptime(request_data['endDate'], "%Y-%m-%d")
        date_range = f"({start_dt.month}.{start_dt.day}-{end_dt.month}.{end_dt.day})"
    else:
        now = datetime.datetime.now()
        yesterday = now - datetime.timedelta(days=1)
        date_range = f"({yesterday.month}.{yesterday.day}-{now.month}.{now.day})"

    path = f"./result-files/{username}"
    if not os.path.exists(path):
        os.makedirs(path)
        logging.debug(f'======================path[{path}]不存在，创建成功========================')
    else:
        logging.debug(f'======================path[{path}]已存在========================')

    file_name = f'公寓学生晚归名单{date_range}.xlsx'
    file_path = f'{path}/{file_name}'
    if os.path.exists(file_path):
        os.remove(file_path)
        logging.debug('======================旧数据删除完成========================')
    new_workbook.save(file_path)
    logging.debug('======================数据导出完成========================')
    return file_path


def set_style(new_sheet1):
    # 设置单元格对齐样式
    align = Alignment(horizontal='center', vertical='center')
    # 设置单元格边框样式
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    # 遍历工作表中的所有单元格并设置样式
    for row in new_sheet1.iter_rows():
        for cell in row:
            cell.alignment = align
            cell.border = border
    # 设置自适应列宽
    for col in new_sheet1.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # 获取列字母
        for cell in col:
            try:  # 需要尝试因为如果单元格为空将会导致错误
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 2
        new_sheet1.column_dimensions[column].width = adjusted_width
